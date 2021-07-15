import sys
import myres_rc
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5 import QtWidgets
import openpyxl
import time

wb = openpyxl.load_workbook('data/worker_list.xlsx', data_only=True)
sheet = wb['worker_list']

list_T = []


def start_six_work(sheet_, sheet2, x, y):
    # 출발빵을 위한 함수
    count = 1
    data = sheet2[chr(83) + '2':chr(83) + '12']
    for row in data:
        for cell in row:
            if cell.value == 0:
                break
            elif y == 1 and count > 7:
                break
            elif y == 2 and count > 4:
                break
            else:
                sheet_.cell(row=x, column=1).value = cell.value
                x = x + 1
                count = count + 1
    return x


def except_return_worker(sheet_, sheet2, index, row_):
    # 복귀자 22시 전에 근무에서 제외하는 함수
    if sheet_.cell(row=row_, column=1).value == row_:
        breaker = True
        for i in range(76, 82):
            data = sheet2[chr(i) + '2':chr(i) + '12']
            for row in data:
                for cell in row:
                    if cell.value == 0:
                        break
                    elif cell.value == list_T[index]:
                        # print(cell.value)
                        index = index + 1
                        if index == len(list_T):
                            index = 0
                        index = except_return_worker(sheet_, sheet2, index, row_)
                        breaker = False
                        break
                    else:
                        continue
                if not breaker:
                    break
            if not breaker:
                break
    return index


def check_worker(name):
    # 입력된 이름이 근무순번에 있는지 확인하는 함수
    for i in range(1, len(list_T) + 1):
        if name == list_T[i - 1]:
            break
        elif i == len(list_T) and name != list_T[i - 1]:
            return "false"
        else:
            continue


def find_slash(num, day, sheet_, wb_):
    # 출력된 근무표 Sheet2 or Sheet4에 근무자 이름 뒤에 붙은 '/'를 제거하는 함수
    check = 0
    for x in range(1, num):
        if x == num - 1:
            check = 1
        if len(sheet_.cell(row=x, column=1).value) == 4:
            list_str = sheet_.cell(row=x, column=1).value.replace("/", "")
            sheet_.cell(row=x, column=1).value = list_str
        else:
            continue
    if check == 1:
        wb_.save('근무표/' + day + '_work_schedule.xlsx')
    else:
        print("fail")


def find_first_worker(w_name):
    # list_T 배열 데이터에서 입력된 첫 근무자를 찾아 그 배열 인덱스 값을 반환하는 함수
    first_worker = w_name
    for i in range(len(list_T)):
        name = list_T[i]
        if name[0:3] == first_worker:
            return i
        else:
            continue


def del_worker():
    # data 폴더 worker_list.xlsx except_list 시트에서 근무에서 제외 할 이름을 list_T와 매칭해 제거하는 함수
    sheet2 = wb['except_list']

    for i in range(65, 71):
        data = sheet2[chr(i) + '2':chr(i) + '12']
        for row in data:
            for cell in row:
                if cell.value == 0:
                    break
                else:
                    # k는 0~56
                    for k in range(len(list_T)):
                        if list_T[k] == cell.value:
                            # print(list_T[k])
                            del list_T[k]
                            break


def find_100below_worker():
    # data 폴더 worker_list.xlsx except_list 시트에서 전입 100일이 안 지난 이름을 찾아 '/'를 추가하는 함수
    sheet2 = wb['except_list']

    for i in range(72, 75):
        data = sheet2[chr(i) + '2':chr(i) + '12']
        for row in data:
            for cell in row:
                if cell.value == 0:
                    break
                else:
                    for k in range(len(list_T)):
                        if list_T[k] == cell.value:
                            list_T[k] = list_T[k] + '/'
                            # print(list_T[k])
                            break


class Weekday:
    # 평일 근무표 클래스
    @staticmethod
    def change_100below(now):
        # 전입 100일 미만 대원 혼자 서는 근무지에서 제외 시키는 함수
        wb_ = openpyxl.load_workbook('근무표/' + now + '_work_schedule.xlsx')
        sheet_ = wb_['Sheet2']
        except_list = [4, 20, 21, 32, 49, 60, 70, 78, 91, 98, 103, 108, 113]

        for x in except_list:
            if len(sheet_.cell(row=x, column=1).value) == 4:
                if len(sheet_.cell(row=x - 1, column=1).value) == 4:
                    if len(sheet_.cell(row=x - 2, column=1).value) == 4:
                        if len(sheet_.cell(row=x - 3, column=1).value) == 4:
                            if len(sheet_.cell(row=x - 4, column=1).value) == 4:
                                name = sheet_.cell(row=x - 5, column=1).value
                                sheet_.cell(row=x - 5, column=1).value = sheet_.cell(row=x - 4, column=1).value
                                sheet_.cell(row=x - 4, column=1).value = name
                            name = sheet_.cell(row=x - 4, column=1).value
                            sheet_.cell(row=x - 4, column=1).value = sheet_.cell(row=x - 3, column=1).value
                            sheet_.cell(row=x - 3, column=1).value = name
                        name = sheet_.cell(row=x - 3, column=1).value
                        sheet_.cell(row=x - 3, column=1).value = sheet_.cell(row=x - 2, column=1).value
                        sheet_.cell(row=x - 2, column=1).value = name
                    name = sheet_.cell(row=x - 2, column=1).value
                    sheet_.cell(row=x - 2, column=1).value = sheet_.cell(row=x - 1, column=1).value
                    sheet_.cell(row=x - 1, column=1).value = name
                # print(sheet_.cell(row=x, column=1).value)
                name = sheet_.cell(row=x - 1, column=1).value
                sheet_.cell(row=x - 1, column=1).value = sheet_.cell(row=x, column=1).value
                sheet_.cell(row=x, column=1).value = name
            else:
                continue

        wb_.save('근무표/' + now + '_work_schedule.xlsx')

    @staticmethod
    def data_set():
        # 근무 순번을 list_T로 데이터화하는 함수
        for i in range(65, 71):
            data = sheet[chr(i) + '2':chr(i) + '12']
            for row in data:
                for cell in row:
                    if cell.value == 0:
                        break
                    else:
                        # print(cell.value)
                        list_T.append(cell.value)

    @staticmethod
    def print_data(name):
        # 처리된 데이터를 순번에 맞게 배치하고 새 엑셀 파일 생성하는 함수
        wb_ = openpyxl.load_workbook('data/basic_work_schedule.xlsx')
        sheet_ = wb_['Sheet2']
        sheet2 = wb['except_list']
        x = 1
        index = find_first_worker(name)
        if index is None:
            return
        
        # 출발빵을 위한 함수
        x = start_six_work(sheet_, sheet2, x, 1)

        while True:
            if sheet_.cell(row=113, column=1).value != 113:
                break
            if index == len(list_T):
                index = 0
            # 22시 부터 근무 가능한 대원 22시 전 근무에서 빼는 if문, 함수로 구현해야함
            index = except_return_worker(sheet_, sheet2, index, 93)

            sheet_.cell(row=x, column=1).value = list_T[index]
            x = x + 1
            index = index + 1

        now = time.strftime('%Y%m%d-%H%M%S', time.localtime(time.time()))
        wb_.save('근무표/' + now + '_work_schedule.xlsx')
        return now


class Weekend:
    # 주말 근무표 클래스
    @staticmethod
    def change_100below(now):
        # 전입 100일 미만 대원 혼자 서는 근무지에서 제외 시키는 함수 (주말용)
        wb_ = openpyxl.load_workbook('근무표/' + now + '_work_schedule.xlsx')
        sheet_ = wb_['Sheet4']
        except_list = [4, 5, 9, 10, 14, 18, 19, 23, 24, 28, 29, 33, 34, 38, 39, 44, 45, 50, 51, 56, 57, 62]

        for x in except_list:
            if len(sheet_.cell(row=x, column=1).value) == 4:
                if len(sheet_.cell(row=x - 1, column=1).value) == 4:
                    if len(sheet_.cell(row=x - 2, column=1).value) == 4:
                        if len(sheet_.cell(row=x - 3, column=1).value) == 4:
                            if len(sheet_.cell(row=x - 4, column=1).value) == 4:
                                name = sheet_.cell(row=x - 5, column=1).value
                                sheet_.cell(row=x - 5, column=1).value = sheet_.cell(row=x - 4, column=1).value
                                sheet_.cell(row=x - 4, column=1).value = name
                            name = sheet_.cell(row=x - 4, column=1).value
                            sheet_.cell(row=x - 4, column=1).value = sheet_.cell(row=x - 3, column=1).value
                            sheet_.cell(row=x - 3, column=1).value = name
                        name = sheet_.cell(row=x - 3, column=1).value
                        sheet_.cell(row=x - 3, column=1).value = sheet_.cell(row=x - 2, column=1).value
                        sheet_.cell(row=x - 2, column=1).value = name
                    name = sheet_.cell(row=x - 2, column=1).value
                    sheet_.cell(row=x - 2, column=1).value = sheet_.cell(row=x - 1, column=1).value
                    sheet_.cell(row=x - 1, column=1).value = name
                # print(sheet_.cell(row=x, column=1).value)
                name = sheet_.cell(row=x - 1, column=1).value
                sheet_.cell(row=x - 1, column=1).value = sheet_.cell(row=x, column=1).value
                sheet_.cell(row=x, column=1).value = name
            else:
                continue

        wb_.save('근무표/' + now + '_work_schedule.xlsx')

    @staticmethod
    def data_set():
        # 근무 순번을 list_T로 데이터화하는 함수 (주말용)
        for i in range(65, 71):
            data = sheet[chr(i) + '2':chr(i) + '12']
            for row in data:
                for cell in row:
                    if cell.value == 0:
                        break
                    else:
                        # print(cell.value)
                        list_T.append(cell.value)

    @staticmethod
    def print_data(name):
        # 처리된 데이터를 순번에 맞게 배치하고 새 엑셀 파일 생성하는 함수 (주말용)
        wb_ = openpyxl.load_workbook('data/basic_work_schedule.xlsx')
        sheet_ = wb_['Sheet4']
        sheet2 = wb['except_list']
        x = 1
        index = find_first_worker(name)

        if index is None:
            return

        # 출발빵을 위한 함수
        x = start_six_work(sheet_, sheet2, x, 2)

        while True:
            if sheet_.cell(row=62, column=1).value != 62:
                break
            if index == len(list_T):
                index = 0

            # 22시 부터 근무 가능한 대원 22시 전 근무에서 빼는 if문, 함수로 구현해야함
            index = except_return_worker(sheet_, sheet2, index, 39)

            sheet_.cell(row=x, column=1).value = list_T[index]
            x = x + 1
            index = index + 1

        now = time.strftime('%Y%m%d-%H%M%S', time.localtime(time.time()))
        wb_.save('근무표/' + now + '_work_schedule.xlsx')
        return now


class Holiday:
    # 공휴일 근무표 클래스
    @staticmethod
    def change_100below(now):
        # 전입 100일 미만 대원 혼자 서는 근무지에서 제외 시키는 함수 (주말용)
        wb_ = openpyxl.load_workbook('근무표/' + now + '_work_schedule.xlsx')
        sheet_ = wb_['Sheet6']
        except_list = [4, 5, 9, 13, 17, 21, 25, 29, 33, 34, 38, 39, 43, 44, 45, 49, 50, 51, 55, 56]

        for x in except_list:
            if len(sheet_.cell(row=x, column=1).value) == 4:
                if len(sheet_.cell(row=x - 1, column=1).value) == 4:
                    if len(sheet_.cell(row=x - 2, column=1).value) == 4:
                        if len(sheet_.cell(row=x - 3, column=1).value) == 4:
                            if len(sheet_.cell(row=x - 4, column=1).value) == 4:
                                name = sheet_.cell(row=x - 5, column=1).value
                                sheet_.cell(row=x - 5, column=1).value = sheet_.cell(row=x - 4, column=1).value
                                sheet_.cell(row=x - 4, column=1).value = name
                            name = sheet_.cell(row=x - 4, column=1).value
                            sheet_.cell(row=x - 4, column=1).value = sheet_.cell(row=x - 3, column=1).value
                            sheet_.cell(row=x - 3, column=1).value = name
                        name = sheet_.cell(row=x - 3, column=1).value
                        sheet_.cell(row=x - 3, column=1).value = sheet_.cell(row=x - 2, column=1).value
                        sheet_.cell(row=x - 2, column=1).value = name
                    name = sheet_.cell(row=x - 2, column=1).value
                    sheet_.cell(row=x - 2, column=1).value = sheet_.cell(row=x - 1, column=1).value
                    sheet_.cell(row=x - 1, column=1).value = name
                # print(sheet_.cell(row=x, column=1).value)
                name = sheet_.cell(row=x - 1, column=1).value
                sheet_.cell(row=x - 1, column=1).value = sheet_.cell(row=x, column=1).value
                sheet_.cell(row=x, column=1).value = name
            else:
                continue

        wb_.save('근무표/' + now + '_work_schedule.xlsx')

    @staticmethod
    def data_set():
        # 근무 순번을 list_T로 데이터화하는 함수 (주말용)
        for i in range(65, 71):
            data = sheet[chr(i) + '2':chr(i) + '12']
            for row in data:
                for cell in row:
                    if cell.value == 0:
                        break
                    else:
                        list_T.append(cell.value)

    @staticmethod
    def print_data(name):
        # 처리된 데이터를 순번에 맞게 배치하고 새 엑셀 파일 생성하는 함수 (주말용)
        wb_ = openpyxl.load_workbook('data/basic_work_schedule.xlsx')
        sheet_ = wb_['Sheet6']
        sheet2 = wb['except_list']
        x = 1
        index = find_first_worker(name)

        if index is None:
            return

        # 출발빵을 위한 함수
        x = start_six_work(sheet_, sheet2, x, 2)

        while True:
            if sheet_.cell(row=56, column=1).value != 56:
                break
            if index == len(list_T):
                index = 0

            # 22시 부터 근무 가능한 대원 22시 전 근무에서 빼는 if문, 함수로 구현해야함
            index = except_return_worker(sheet_, sheet2, index, 37)

            sheet_.cell(row=x, column=1).value = list_T[index]
            x = x + 1
            index = index + 1

        now = time.strftime('%Y%m%d-%H%M%S', time.localtime(time.time()))
        wb_.save('근무표/' + now + '_work_schedule.xlsx')
        return now


# UI파일 연결
# 단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form_class = uic.loadUiType("ui/home.ui")[0]


# 화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.bt_weekday.clicked.connect(self.groupboxRadFunction_1)
        self.bt_weekend.clicked.connect(self.groupboxRadFunction_2)
        self.bt_holiday.clicked.connect(self.groupboxRadFunction_3)
        self.bt_new.clicked.connect(self.get_file)
        self.bt_time.clicked.connect(self.del_slash)

    sel = 3
    w_day = 3

    # def showMessageBox(self):
    #     msgobj1 = QtWidgets.QMessageBox(self)
    #     msgobj1.question(self, 'MessageBox title', 'Here comes message',
    #                      QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)

    def groupboxRadFunction_1(self):
        self.sel = 0
        self.w_day = 0
        print("Chekced weekday")

    def groupboxRadFunction_2(self):
        self.sel = 1
        self.w_day = 1
        print("Checked weekend")

    def groupboxRadFunction_3(self):
        self.sel = 2
        self.w_day = 2
        print("Checked holiday")

    def get_file(self):

        if not self.bt_weekday.isChecked() and not self.bt_weekend.isChecked() and not self.bt_holiday.isChecked():
            msgobj1 = QtWidgets.QMessageBox(self)
            msgobj1.question(self, '알림', '근무표 유형을 선택하세요.',
                             QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            return

        name = self.li_worker.text()

        if (self.bt_weekday.isChecked() or self.bt_weekend.isChecked() or self.bt_holiday.isChecked()) and len(name) >= 5:
            msgobj1 = QtWidgets.QMessageBox(self)
            msgobj1.question(self, '알림', '알맞는 이름을 입력하세요.',
                             QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            return

        if self.sel == 0 and self.bt_weekday.isChecked:
            self.progressBar.setValue(0)
            today = Weekday
            self.progressBar.setValue(10)
            today.data_set()
            self.progressBar.setValue(30)
            del_worker()
            self.progressBar.setValue(50)
            find_100below_worker()
            self.progressBar.setValue(70)
            now = today.print_data(name)
            self.progressBar.setValue(80)
            if now is None:
                msgobj1 = QtWidgets.QMessageBox(self)
                msgobj1.question(self, '알림', '알맞는 이름을 입력하세요.',
                                 QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                return
            for i in range(5):
                today.change_100below(now)
            self.progressBar.setValue(100)
        elif self.sel == 1 and self.bt_weekend.isChecked:
            self.progressBar.setValue(0)
            today = Weekend
            self.progressBar.setValue(10)
            today.data_set()
            self.progressBar.setValue(30)
            del_worker()
            self.progressBar.setValue(50)
            find_100below_worker()
            self.progressBar.setValue(70)
            now = today.print_data(name)
            self.progressBar.setValue(80)
            if now is None:
                msgobj1 = QtWidgets.QMessageBox(self)
                msgobj1.question(self, '알림', '알맞는 이름을 입력하세요.',
                                 QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                return
            for i in range(5):
                today.change_100below(now)
            self.progressBar.setValue(100)
        elif self.sel == 2 and self.bt_holiday.isChecked:
            self.progressBar.setValue(0)
            today = Holiday
            self.progressBar.setValue(10)
            today.data_set()
            self.progressBar.setValue(30)
            del_worker()
            self.progressBar.setValue(50)
            find_100below_worker()
            self.progressBar.setValue(70)
            now = today.print_data(name)
            self.progressBar.setValue(80)
            if now is None:
                msgobj1 = QtWidgets.QMessageBox(self)
                msgobj1.question(self, '알림', '알맞는 이름을 입력하세요.',
                                 QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                return
            for i in range(5):
                today.change_100below(now)
            self.progressBar.setValue(100)

    def del_slash(self):

        if not self.bt_weekday.isChecked() and not self.bt_weekend.isChecked() and not self.bt_holiday.isChecked():
            msgobj1 = QtWidgets.QMessageBox(self)
            msgobj1.question(self, '알림', '근무표 유형을 선택하세요.',
                             QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            return

        if len(self.li_time.text()) < 15 or len(self.li_time.text()) > 15:
            msgobj1 = QtWidgets.QMessageBox(self)
            msgobj1.question(self, '알림', '알맞는 날짜-시간을 입력하세요.',
                             QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            return

        self.progressBar_2.setValue(0)
        day = self.li_time.text()
        wb_ = openpyxl.load_workbook('근무표/' + day + '_work_schedule.xlsx')
        self.progressBar_2.setValue(20)
        if self.w_day == 0:
            sheet_ = wb_['Sheet2']
            if sheet_.cell(row=1, column=1).value == 1:
                msgobj1 = QtWidgets.QMessageBox(self)
                msgobj1.question(self, '알림', '근무표 유형이 틀렸습니다.',
                                 QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                return
            num = 114
        elif self.w_day == 1:
            sheet_ = wb_['Sheet4']
            if sheet_.cell(row=1, column=1).value == 1:
                msgobj1 = QtWidgets.QMessageBox(self)
                msgobj1.question(self, '알림', '근무표 유형이 틀렸습니다.',
                                 QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                return
            num = 63
        elif self.w_day == 2:
            sheet_ = wb_['Sheet6']
            if sheet_.cell(row=1, column=1).value == 1:
                msgobj1 = QtWidgets.QMessageBox(self)
                msgobj1.question(self, '알림', '근무표 유형이 틀렸습니다.',
                                 QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                return
            num = 57
        else:
            print("fail")
        self.progressBar_2.setValue(50)
        find_slash(num, day, sheet_, wb_)
        self.progressBar_2.setValue(100)


if __name__ == "__main__":
    # QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv)

    # WindowClass의 인스턴스 생성
    myWindow = WindowClass()

    # 프로그램 화면을 보여주는 코드
    myWindow.show()

    # 프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()
